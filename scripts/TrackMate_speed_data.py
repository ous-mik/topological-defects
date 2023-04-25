import matplotlib.pyplot as plt
import pandas as pd
import numpy as np
import os
import openpyxl as openpyxl

def extract_file_names_from_directory(directory):
    file_names = []
    for subdir, dirs, files in os.walk(directory):
        for file in files:
            file_names.append(os.path.join(file))
    return file_names

def average_single_site(file):
    df = pd.read_csv(path + file, usecols=["SPEED", "EDGE_TIME"], low_memory=False, encoding='latin-1')
    column = df["EDGE_TIME"]
    column = column[3:].to_numpy(dtype=float)
    range = np.arange(0.5, column.max() + 1, 1)
    average_list = []
    for i in range:
        time_point = df.loc[df['EDGE_TIME'] == str(i)]
        speed = time_point["SPEED"]
        speed = speed[3:].to_numpy(dtype=float)
        average = np.average(speed)
        average_list.append(average)
    return average_list

def get_column_list(list, ark):
    column_list = []
    for name in list:
        for i in range(0, ark.max_column):
            if name in str(ark.cell(row = 1, column = i + 1).value):
                column_list.append(i+1)
    return(column_list)

def get_column(column_number, ark):
    column=np.zeros(ark.max_row)
    for i in range(0, ark.max_row-2):
        column[i] = ark.cell(column = column_number, row = i +2).value
    return(column)

def get_row_average(column_list, ark, row_number):
    row_numbers  = []
    for i in column_list:
        row_numbers.append(ark.cell(column = i, row = row_number).value)
    average = np.average((row_numbers))
    return(average)

def X_axis(interval, ark):
    X = np.zeros(ark.max_row)
    for i in range(0, ark.max_row):
        X[i] = (i*interval)/60
    return(X)

#Define the path where the files are stored.
path = "......"

# Get the file names
file_names = extract_file_names_from_directory(directory = path)

print(file_names)

List_of_data = []
header = []

for file in file_names:
    data=average_single_site(file)
    List_of_data.append(data)
    name = file[:-9]
    header.append(name)

#Create a dataframe with all the data
df = pd.DataFrame(data=List_of_data, index=header).T

print(df)

# Save to Excel
df.to_excel('Data.xlsx', index=False, header=header)

#Plot data from excel file
wb = openpyxl.load_workbook("Data.xlsx", data_only=True)
ark = wb.active

# Put names of columns to be plotted here
plot_list = []

# For each line in the graph put a label here
label_list = []

average_list = []
std_list = []

for group in plot_list:
    column_list=get_column_list(list=group, ark=ark)
    columns=[]
    for i in column_list:
        column=get_colun(column_number=i, ark=ark)
        columns.append(column)
    average=(np.average(columns, axis=0))*0.7792764     #Multiply with pixelsize (µm/pixel)
    Average=(average/4)*60                              #Divide with time interval and multiply with 60 to get values in µm/h
    std=(np.std(columns, axis=0))*0.7792764             #Multiply with pixelsize (µm/pixel)
    STD=(std/4)*60                                      #Divide with time interval and multiply with 60 to get values in µm/h
    average_list.append(Average)
    std_list.append(STD)

color_list = []

x = X_axis(interval = 4, ark = ark)
X = x + 1

fig, ax = plt.subplots(ncols=1, nrows=1, figsize = (4, 4))

for average, std, label, color in zip(average_list, std_list, label_list, color_list):
    plt.plot(X[:], average[:], alpha = 1, linewidth=0.5, label=label, color=color)
    plt.fill_between(X[:], average[:]-std[:], average[:]+std[:], color=color, alpha=0.2, edgecolor=None)

plt.legend(loc='best', prop={'size': 8})
plt.ylim(0, 80)
plt.xlabel("Time after stimulation (h)", fontsize=14)
plt.ylabel('Average cell speed (µm/h)', fontsize=14)
fig.tight_layout()
plt.savefig("Fig1.png", dpi=300)

plt.show()
